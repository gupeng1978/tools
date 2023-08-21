from openpyxl import Workbook


def merge_cells_for_header(worksheet, excel_row_index):
    row_values = [cell.value for cell in worksheet[excel_row_index]]
    row_len = len(row_values)
    combine_start = 0    
    while combine_start < row_len:
        combine_end = combine_start + 1
        while combine_end < row_len:
            if row_values[combine_end] != row_values[combine_start]:
                break
            combine_end += 1
        
        # combine if possible
        if combine_end != combine_start + 1:
            worksheet.merge_cells(start_row=excel_row_index, start_column=combine_start+1,
                                    end_row=excel_row_index, end_column=combine_end)
        combine_start = combine_end
        


def merge_cells_for_records(worksheet, excel_row_start, excel_row_end, sort_key_row_index):
    
    # 遍历每一列
    allowed_seg_table = [{'start':excel_row_start, 'end':excel_row_end}]
    next_allowed_seg_table = []
    
    def get_segments_from_list(req_seg):
        allowed_segments = []
        
        # 遍历允许的分段表
        for allowed_seg in allowed_seg_table:
            # 检查请求分段与允许分段的交集
            start = max(req_seg['start'], allowed_seg['start'])
            end = min(req_seg['end'], allowed_seg['end'])
            
            # 如果存在交集，则添加到允许的分段列表中
            if start < end:
                allowed_segments.append({'start': start, 'end': end})
        
        return allowed_segments
    
    def try_merged(start_row, end_row, column) :
        allowed_seg = get_segments_from_list({'start':start_row, 'end':end_row})
        if allowed_seg :
            for seg in allowed_seg:
                worksheet.merge_cells(
                    start_row=seg['start'], start_column=column,
                    end_row=seg['end'], end_column=column
                )
            next_allowed_seg_table.append(seg)        
        pass
    
    for col in worksheet.iter_cols(min_row=excel_row_start, max_row=excel_row_end):
        
        # 只合并sort_key_row_index所在的列
        if col[0].column - 1 not in sort_key_row_index:
            continue
        next_allowed_seg_table = []
        
        col_number = col[0].column
        combine_start = excel_row_start
        previous_value = None
        # 遍历列中的每一行
        for row_index, cell in enumerate(col, start=excel_row_start):
            if previous_value is None:
                previous_value = cell.value
                continue
            
            # 如果当前值与前一个值相同，继续遍历
            if cell.value == previous_value:
                continue
            
            # 如果当前值与前一个值不同，检查是否需要合并
            if row_index - combine_start > 1:
                try_merged(combine_start, row_index - 1, cell.column)
            
            # 更新合并的起始行和前一个值
            combine_start = row_index
            previous_value = cell.value
        
        # 检查最后一组单元格是否需要合并
        if excel_row_end - combine_start > 0:
            try_merged(combine_start, excel_row_end, cell.column)
        
        # 更新下一次迭代允许的分段    
        allowed_seg_table = next_allowed_seg_table
                


def extend_record_by_keys(record_dict, keys_list):
    result = [record_dict.get(key, None) for key in keys_list]
    return result
            

class Header:
    def __init__(self):
        self.headers = []
        self.aliases = {}
        self.header_col = 0
        self.record_keys = None

    def add(self, header_list):       
        if isinstance(header_list, list) and all(isinstance(h, str) for h in header_list):
            if self.header_col == 0:            
                self.headers.append(header_list)
                self.header_col = len(header_list)
                self.record_keys = header_list
            else:                
                if self.header_col != len(header_list):
                    raise ValueError("多个Header参数长度不匹配")                    
                self.headers.append(header_list)
            
        else:   
            raise ValueError("Header 参数必须是字符串列表")

    def set_active(self, index, sort_keys):
        if isinstance(index, int) and 0 <= index < len(self.headers):
            self.record_keys = self.headers[index].copy()
            self.sort_keys = sort_keys
        else:
            raise ValueError("索引必须是有效的整数")


    def set_alias(self, alias_dict):
        if isinstance(alias_dict, dict) and all(isinstance(k, str) and isinstance(v, str) for k, v in alias_dict.items()):
            self.aliases.update(alias_dict)
        else:
            raise ValueError("字典的键和值必须是字符串")
        
    


class Record:
    def __init__(self):
        self.records = []

    def add_from_str(self, record_str):
        if isinstance(record_str, str):
            self.records.append(record_str)
        else:
            raise ValueError("参数必须是字符串")

    def add_from_dict(self, record_dict):
        if isinstance(record_dict, dict):
            self.records.append(record_dict)
        else:
            raise ValueError("参数必须是字典")
        
    def expand_by_keys(self, record_keys):
        for index, record in enumerate(self.records):
            if isinstance(record, dict):
                # 检查record的所有键是否都在record_keys中
                if not all(key in record_keys for key in record.keys()):
                    raise ValueError("所有的record的keys都必须在record_keys中")
                self.records[index] = [record.get(key, None) for key in record_keys]
            else:
                raise TypeError("记录必须是字典类型")
            
    def sort_by_keys(self, record_keys, sort_keys):
        # 检查sort_keys是否是record_keys的子集
        if not all(key in record_keys for key in sort_keys):
            raise ValueError("sort_keys必须是record_keys的子集")

        # 计算sort_keys在record_keys的索引列表
        sort_indexes = [record_keys.index(key) for key in sort_keys]

        # 根据sort_indexes对self.records进行排序
        self.records.sort(key=lambda record: [record[index] for index in sort_indexes])
        
        return sort_indexes

        
    
            
        


class Table:
    def __init__(self, header, record):
        if not isinstance(header, Header) or not isinstance(record, Record):
            raise TypeError("header and record must be Header and Record objects respectively.")
        self.header = header
        self.record = record
        
        # replace the header name with alias name
        for h in self.header.headers:
            for i, l in enumerate(h):
                if l in header.aliases:
                    h[i] = header.aliases[l]
                    
        # check self.header.record_keys list strings cannot be the same
        if len(set(self.header.record_keys)) != len(self.header.record_keys):
            raise ValueError("用户设置header的record_keys 包含重复字符串")

                    

    def to_excel(self, worksheet):
       
        # 将表头写入工作表
        for header in self.header.headers:
            worksheet.append(header)        
            merge_cells_for_header(worksheet, worksheet.max_row)

        # 将记录写入工作表
        self.record.expand_by_keys(self.header.record_keys)
        sort_key_row_index = self.record.sort_by_keys(self.header.record_keys, self.header.sort_keys)
        record_row_start = worksheet.max_row + 1
        for record in self.record.records:            
            worksheet.append(record)
        record_row_end = worksheet.max_row
        merge_cells_for_records(worksheet, record_row_start, record_row_end, sort_key_row_index)


        # 输出表格数据到控制台
        for row in worksheet.iter_rows(values_only=True):
            print(row)

            

# 示例
header = Header()
header.add(["model", "thread", "step","device_nn", "device_nn", "host_pcie_nn", "host_pcie_nn"])
header.add(["model", "thread", "step","dev_ms",    "dev_fps",   "host_ms",      "host_fps"])
header.set_active(1, ['model','thread'])
header.set_alias({"model": "模型", 
                  "thread": "线程数", 
                  "step": "步骤", 
                  "dev_ms": "时间(ms)", 
                  "dev_fps": "帧率", 
                  "host_ms": "时间(ms)", 
                  "host_fps": "帧率",
                  "device_nn":"device推理",
                  "host_pcie_nn":"Host PCIe推理"})

record = Record()
# record.add_from_str("John, 25")

record.add_from_dict({"model": "resnet50", 
                      "thread": '2', 
                      "step" :  "dclmdlLoadFromFile",
                      "dev_ms" : "2",
                      "host_ms" : "1.59"})

record.add_from_dict({"model": "resnet50", 
                      "thread": '1', 
                      "step" :  "dclmdlLoadFromFile",
                      "dev_ms" : "2.59",
                      "host_ms" : "3.59"})

record.add_from_dict({"model": "resnet50", 
                      "thread": '1', 
                      "step" :  "copyH2D",
                      "host_ms" : "0.5"})

record.add_from_dict({"model": "resnet50", 
                      "thread": '1', 
                      "step" :  "copyD2H",
                      "host_ms" : "1.5"})

record.add_from_dict({"model": "yolov8", 
                      "thread": '16', 
                      "step" :  "copyD2H",
                      "host_ms" : "3.5"})

record.add_from_dict({"model": "resnet50", 
                      "thread": '2', 
                      "step" :  "copyH2D",
                      "host_ms" : "0.5"})

record.add_from_dict({"model": "resnet50", 
                      "thread": '4', 
                      "step" :  "copyD2H",
                      "host_ms" : "3.5"})

table = Table(header, record)

 # 创建一个工作簿和工作表
workbook = Workbook()
worksheet = workbook.active
        
table.to_excel(worksheet)

 # 保存工作簿
workbook.save("output1.xlsx")