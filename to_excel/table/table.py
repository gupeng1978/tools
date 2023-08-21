from openpyxl import Workbook
from table_merge import Table_Merge


class Header:
    def __init__(self):
        self.headers = []
        self.aliases = {}
        self.header_col = 0
        self.record_keys = None

    def add(self, header_list):       
        if isinstance(header_list, list) and all(isinstance(h, str) for h in header_list):
            if self.header_col == 0:            
                self.headers.append(header_list)
                self.header_col = len(header_list)
                self.record_keys = header_list
            else:                
                if self.header_col != len(header_list):
                    raise ValueError("多个Header参数长度不匹配")                    
                self.headers.append(header_list)
            
        else:   
            raise ValueError("Header 参数必须是字符串列表")

    def set_active(self, index, sort_keys):
        if isinstance(index, int) and 0 <= index < len(self.headers):
            self.record_keys = self.headers[index].copy()
            self.sort_keys = sort_keys
        else:
            raise ValueError("索引必须是有效的整数")


    def set_alias(self, alias_dict):
        if isinstance(alias_dict, dict) and all(isinstance(k, str) and isinstance(v, str) for k, v in alias_dict.items()):
            self.aliases.update(alias_dict)
        else:
            raise ValueError("字典的键和值必须是字符串")
        
    def __str__(self):
        header_str = "Headers:\n" + "\n".join(", ".join(header) for header in self.headers)
        alias_str = "Aliases:\n" + "\n".join(f"{k}: {v}" for k, v in self.aliases.items())
        active_header_str = "Active Header:\n" + ", ".join(self.record_keys) if self.record_keys else "None"
        return f"{header_str}\n\n{alias_str}\n\n{active_header_str}"
        
    


class Record:
    def __init__(self):
        self.records = []

    def add_from_str(self, record_str):
        if isinstance(record_str, str):
            self.records.append(record_str)
        else:
            raise ValueError("参数必须是字符串")

    def add_from_dict(self, record_dict):
        if isinstance(record_dict, dict):
            self.records.append(record_dict)
        else:
            raise ValueError("参数必须是字典")
        
    def __str__(self):
        records_str = "\n".join(str(record) for record in self.records)
        return f"Records:\n{records_str}"


class Table:
    def __init__(self, worksheet, header, record):
        if not isinstance(header, Header) or not isinstance(record, Record):
            raise TypeError("header and record must be Header and Record objects respectively.")
        self.__worksheet = worksheet
        self.__header = header
        self.__record = record       
        self.__table_info = {'header' : {'row_start':None, 'row_end':None}, 'record' : {'row_start':None, 'row_end':None}}        
        self.__sort_key_row_index = None
        
        
        # replace the header name with alias name
        for h in self.__header.headers:
            for i, l in enumerate(h):
                if l in header.aliases:
                    h[i] = header.aliases[l]
                    
        # check self.header.record_keys list strings cannot be the same
        if len(set(self.__header.record_keys)) != len(self.__header.record_keys):
            raise ValueError("用户设置header的record_keys 包含重复字符串")
        
        self.__to_excel()

            
    def __expand_and_sort_by_keys(self, record_keys, sort_keys):
        # 检查sort_keys是否是record_keys的子集
        if not all(key in record_keys for key in sort_keys):
            raise ValueError("sort_keys必须是record_keys的子集")
        
        # expand
        for index, record in enumerate(self.__record.records):
            if isinstance(record, dict):
                # 检查record的所有键是否都在record_keys中
                if not all(key in record_keys for key in record.keys()):
                    raise ValueError("所有的record的keys都必须在record_keys中")
                self.__record.records[index] = [record.get(key, None) for key in record_keys]
            else:
                raise TypeError("记录必须是字典类型")

        # sort 
        # 计算sort_keys在record_keys的索引列表
        sort_indexes = [record_keys.index(key) for key in sort_keys]

        # 根据sort_indexes对self.records进行排序
        self.__record.records.sort(key=lambda record: [record[index] for index in sort_indexes])
        
        self.__sort_key_row_index = sort_indexes
    
    
    def merge_cells(self):
        table_merge = Table_Merge(self.__worksheet, self.__table_info, self.__sort_key_row_index)
        table_merge.merge()        
        
    def __str__(self):
        # 获取表格信息中的起始和结束行
        row_start = self.__table_info['header']['row_start']
        row_end = self.__table_info['record']['row_end']

        # 使用列表推导式构建一个包含指定行范围的字符串列表
        rows = ['\t'.join(map(str, row)) for row in self.__worksheet.iter_rows(min_row=row_start, max_row=row_end, values_only=True)]

        # 使用换行符连接所有行，构建一个完整的字符串
        return '\n'.join(rows)    
    

    def __to_excel(self):        
        # 计算表格信息
        worksheet = self.__worksheet
        self.__table_info['header']['row_start'] = worksheet.max_row
        self.__table_info['header']['row_end'] = self.__table_info['header']['row_start'] + len(self.__header.headers) - 1
        self.__table_info['record']['row_start'] = self.__table_info['header']['row_end'] + 1
        self.__table_info['record']['row_end'] = self.__table_info['record']['row_start'] + len(self.__record.records) - 1
       
        # 将表头写入工作表        
        for header in self.__header.headers:
            worksheet.append(header)        

        # 将记录写入工作表
        self.__expand_and_sort_by_keys(self.__header.record_keys, self.__header.sort_keys)
        record_row_start = worksheet.max_row + 1
        for record in self.__record.records:            
            worksheet.append(record)
        record_row_end = worksheet.max_row


        

            

# # 示例
# header = Header()
# header.add(["model", "thread", "step","device_nn", "device_nn", "host_pcie_nn", "host_pcie_nn"])
# header.add(["model", "thread", "step","dev_ms",    "dev_fps",   "host_ms",      "host_fps"])
# header.set_active(1, ['model','thread'])
# header.set_alias({"model": "模型", 
#                   "thread": "线程数", 
#                   "step": "步骤", 
#                   "dev_ms": "时间(ms)", 
#                   "dev_fps": "帧率", 
#                   "host_ms": "时间(ms)", 
#                   "host_fps": "帧率",
#                   "device_nn":"device推理",
#                   "host_pcie_nn":"Host PCIe推理"})

# print(header)

# record = Record()
# # record.add_from_str("John, 25")

# record.add_from_dict({"model": "resnet50", 
#                       "thread": '2', 
#                       "step" :  "dclmdlLoadFromFile",
#                       "dev_ms" : "2",
#                       "host_ms" : "1.59"})

# record.add_from_dict({"model": "resnet50", 
#                       "thread": '1', 
#                       "step" :  "dclmdlLoadFromFile",
#                       "dev_ms" : "2.59",
#                       "host_ms" : "3.59"})

# record.add_from_dict({"model": "resnet50", 
#                       "thread": '1', 
#                       "step" :  "copyH2D",
#                       "host_ms" : "0.5"})

# record.add_from_dict({"model": "resnet50", 
#                       "thread": '1', 
#                       "step" :  "copyD2H",
#                       "host_ms" : "1.5"})

# record.add_from_dict({"model": "yolov8", 
#                       "thread": '16', 
#                       "step" :  "copyD2H",
#                       "host_ms" : "3.5"})

# record.add_from_dict({"model": "resnet50", 
#                       "thread": '2', 
#                       "step" :  "copyH2D",
#                       "host_ms" : "0.5"})

# record.add_from_dict({"model": "resnet50", 
#                       "thread": '4', 
#                       "step" :  "copyD2H",
#                       "host_ms" : "3.5"})

# print(record)

#  # 创建一个工作簿和工作表
# workbook = Workbook()
# worksheet = workbook.active

# table = Table(worksheet, header, record)
# table.merge_cells()
        
# print(table)

#  # 保存工作簿
# workbook.save("output1.xlsx")

