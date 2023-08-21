from openpyxl import Workbook
import unittest


class Table_Merge:
    def __init__(self, worksheet, table_info, sort_key_row_index):
        self.__worksheet = worksheet
        self.__table_info = table_info
        self.__sort_key_row_index = sort_key_row_index
        pass
    
    def __merge_cells_by_row(self, row_start, row_end):
        worksheet = self.__worksheet
        for row_cells in worksheet.iter_rows(min_row=row_start, max_row=row_end):
            start_col = None
            value_to_merge = None
            for col, cell in enumerate(row_cells, start=1):
                if cell.value == value_to_merge:
                    continue
                if start_col:
                    worksheet.merge_cells(start_row=cell.row, start_column=start_col, end_row=cell.row, end_column=col - 1)
                start_col = col
                value_to_merge = cell.value
            if start_col:
                worksheet.merge_cells(start_row=cell.row, start_column=start_col, end_row=cell.row, end_column=col)
                
    def __merge_cells_by_col(self, row_start, row_end):
        worksheet = self.__worksheet
        for col_cells in worksheet.iter_cols(min_row=row_start, max_row=row_end):
            start_row = None
            value_to_merge = None
            for row, cell in enumerate(col_cells, start=row_start):
                if cell.value == value_to_merge:
                    continue
                if start_row:
                    worksheet.merge_cells(start_row=start_row, start_column=cell.column, end_row=row - 1, end_column=cell.column)
                start_row = row
                value_to_merge = cell.value
            if start_row:
                worksheet.merge_cells(start_row=start_row, start_column=cell.column, end_row=row, end_column=cell.column)
    
    
    def __merge_cells_for_records(self, excel_row_start, excel_row_end):
        worksheet = self.__worksheet
        sort_key_row_index = self.__sort_key_row_index
        # 遍历每一列
        allowed_seg_table = [{'start':excel_row_start, 'end':excel_row_end}]
        next_allowed_seg_table = []
        
        def get_segments_from_list(req_seg):
            allowed_segments = []
            
            # 遍历允许的分段表
            for allowed_seg in allowed_seg_table:
                # 检查请求分段与允许分段的交集
                start = max(req_seg['start'], allowed_seg['start'])
                end = min(req_seg['end'], allowed_seg['end'])
                
                # 如果存在交集，则添加到允许的分段列表中
                if start < end:
                    allowed_segments.append({'start': start, 'end': end})
            
            return allowed_segments
        
        def try_merged(start_row, end_row, column) :
            allowed_seg = get_segments_from_list({'start':start_row, 'end':end_row})
            if allowed_seg :
                for seg in allowed_seg:
                    worksheet.merge_cells(
                        start_row=seg['start'], start_column=column,
                        end_row=seg['end'], end_column=column
                    )
                next_allowed_seg_table.append(seg)        
            pass
        
        for col in worksheet.iter_cols(min_row=excel_row_start, max_row=excel_row_end):
            
            # 只合并sort_key_row_index所在的列
            if col[0].column - 1 not in sort_key_row_index:
                continue
            next_allowed_seg_table = []
            
            col_number = col[0].column
            combine_start = excel_row_start
            previous_value = None
            # 遍历列中的每一行
            for row_index, cell in enumerate(col, start=excel_row_start):
                if previous_value is None:
                    previous_value = cell.value
                    continue
                
                # 如果当前值与前一个值相同，继续遍历
                if cell.value == previous_value:
                    continue
                
                # 如果当前值与前一个值不同，检查是否需要合并
                if row_index - combine_start > 1:
                    try_merged(combine_start, row_index - 1, cell.column)
                
                # 更新合并的起始行和前一个值
                combine_start = row_index
                previous_value = cell.value
            
            # 检查最后一组单元格是否需要合并
            if excel_row_end - combine_start > 0:
                try_merged(combine_start, excel_row_end, cell.column)
            
            # 更新下一次迭代允许的分段    
            allowed_seg_table = next_allowed_seg_table
    
    
    def merge(self):        
        if 'header' in self.__table_info:
            header_info = self.__table_info['header']
            self.__merge_cells_by_row(header_info['row_start'], header_info['row_end'])
            self.__merge_cells_by_col(header_info['row_start'], header_info['row_end'])
                
       
        if 'record' in self.__table_info:
            record_info = self.__table_info['record']
            self.__merge_cells_for_records(record_info['row_start'], record_info['row_end'])
        
 