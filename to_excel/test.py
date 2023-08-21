from openpyxl import Workbook


def merge_cells_for_records(worksheet, excel_row_start, excel_row_end, sort_key_row_index):
    
    # 遍历每一列
    allowed_seg_table = [{'start':excel_row_start, 'end':excel_row_end}]
    next_allowed_seg_table = []
    
    def get_segments_from_list(req_seg):
        allowed_segments = []
        
        # 遍历允许的分段表
        for allowed_seg in allowed_seg_table:
            # 检查请求分段与允许分段的交集
            start = max(req_seg['start'], allowed_seg['start'])
            end = min(req_seg['end'], allowed_seg['end'])
            
            # 如果存在交集，则添加到允许的分段列表中
            if start < end:
                allowed_segments.append({'start': start, 'end': end})
        
        return allowed_segments
    
    def try_merged(start_row, end_row, column) :
        allowed_seg = get_segments_from_list({'start':start_row, 'end':end_row})
        if allowed_seg :
            for seg in allowed_seg:
                worksheet.merge_cells(
                    start_row=seg['start'], start_column=column,
                    end_row=seg['end'], end_column=column
                )
            next_allowed_seg_table.append(seg)        
        pass
    
    for col in worksheet.iter_cols(min_row=excel_row_start, max_row=excel_row_end):
        
        # 只合并sort_key_row_index所在的列
        if col[0].column - 1 not in sort_key_row_index:
            continue
        next_allowed_seg_table = []
        
        col_number = col[0].column
        combine_start = excel_row_start
        previous_value = None
        # 遍历列中的每一行
        for row_index, cell in enumerate(col, start=excel_row_start):
            if previous_value is None:
                previous_value = cell.value
                continue
            
            # 如果当前值与前一个值相同，继续遍历
            if cell.value == previous_value:
                continue
            
            # 如果当前值与前一个值不同，检查是否需要合并
            if row_index - combine_start > 1:
                try_merged(combine_start, row_index - 1, cell.column)
            
            # 更新合并的起始行和前一个值
            combine_start = row_index
            previous_value = cell.value
        
        # 检查最后一组单元格是否需要合并
        if excel_row_end - combine_start > 0:
            try_merged(combine_start, excel_row_end, cell.column)
        
        # 更新下一次迭代允许的分段    
        allowed_seg_table = next_allowed_seg_table

# 示例
workbook = Workbook()
worksheet = workbook.active
worksheet.append(['a', 'b', 'b', 'c'])
worksheet.append(['a', 'b', 'c', 'c'])
worksheet.append(['a', 'x', 'c', 'c'])
worksheet.append(['a', 'x', 'c', 'd'])
worksheet.append(['a', 'x', 'c', 'd'])

merge_cells_for_records(worksheet, 1, 4, [0,1,2])

 # 输出表格数据到控制台
for row in worksheet.iter_rows(values_only=True):
    print(row)
workbook.save("output3.xlsx")

