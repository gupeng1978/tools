from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .table import Header, Record, Table
import os
import yaml


TEMPLATE_CFG_FILE = os.path.join(os.path.dirname(__file__), '../data/template.yaml')

class Sheet:
    def __init__(self, worksheet, name='Sheet1'):
        if not isinstance(worksheet, Worksheet):
            raise TypeError("worksheet must be an instance of openpyxl.worksheet.worksheet.Worksheet")

        self.__sheet = worksheet
        self.__name = name


    def __str__(self):
        # 添加工作表名称
        sheet_name_info = f"------- Sheet Name: {self.__name}"

        # 使用列表推导式构建一个包含所有行的字符串列表
        rows = ['\t'.join(map(str, row)) for row in self.__sheet.iter_rows(values_only=True)]

        # 添加合并单元格的信息
        merged_cells_info = "Merged Cells:"
        for merged_range in self.__sheet.merged_cells.ranges:
            merged_cells_info += f"\n{merged_range.coord}"

        # 使用换行符连接工作表名称、所有行和合并单元格的信息，构建一个完整的字符串
        return sheet_name_info + '\n' + '\n'.join(rows) + '\n' + merged_cells_info



def get_default_config(output_file_path):
    # 检查并创建输出目录（如果不存在）
    output_directory = os.path.dirname(output_file_path)
    if output_directory and not os.path.exists(output_directory):
        os.makedirs(output_directory)
        
    with open(TEMPLATE_CFG_FILE, 'r', encoding='utf-8') as template_file:        
        with open(output_file_path, 'w', encoding='utf-8') as output_file:            
            for line in template_file:
                output_file.write(line)
                


def gen_excel_table(config_file):
    config = {}
    excel_path = None
    sheet_tags = None
    # 检查文件是否存在
    if not os.path.isfile(config_file):
        raise ValueError(f"File {config_file} does not exist.")

    # 尝试加载YAML内容以检查其有效性
    try:
        with open(config_file, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
    except yaml.YAMLError:
        raise ValueError(f"File {config_file} is not a valid YAML file.")
    
    
    # excel_path处理
    excel_path = config.get('excel_path')
    if not excel_path or not excel_path.endswith('.xlsx'):
        raise ValueError(f"excel_path must be a .xlsx file. Got: {excel_path}")
    if os.path.exists(excel_path) and not os.access(excel_path, os.W_OK):
        raise ValueError(f"excel_path is not writable. Got: {excel_path}")
    
    # 创建一个工作簿和工作表
    workbook = Workbook()
    
    # 删除缺省sheet
    workbook.remove(workbook.active)

    
    # sheet处理
    sheets = config.get('sheets')
    if not isinstance(sheets, dict) or not isinstance(sheets.get('tag'), list) or not all(isinstance(item, str) for item in sheets.get('tag', [])):
        raise ValueError("sheets must be a dictionary containing a 'tag' key with a list of strings.")
    
    
    excel_sheets = {}
    for sheet_tag in sheets['tag']:        
        excel_sheets[sheet_tag] = workbook.create_sheet(title=sheet_tag)            
        
    
    
    # table处理    
    tables = config.get('tables', [])
    for table in tables:
        if 'name' not in table:
            raise ValueError(f"{config_file} failed, no table name")
        
        # table 的sheet tag处理
        if 'sheet_tag' not in table:
            raise ValueError(f"{config_file} failed,  no table sheet tag")
        
        if table['sheet_tag'] not in sheets['tag']:
            raise ValueError(f"{config_file} failed,  table sheet is not a valid sheet name, sheets = {table['sheet_tag']}")        
        table_sheet = excel_sheets[table['sheet_tag']]
        
        #header 处理
        header = Header()
        if 'head-key' not in table:
            raise ValueError(f"{config_file} failed,  no table head-key")
        
        if 'head-0' in table:
            header.add(table['head-0'])
        else:
            raise ValueError(f"{config_file} failed,  no table head-0")
        
        if 'head-1' in table:
            header.add(table['head-1'])
            header.set_active(1, table['head-key'])
        else:
            header.set_active(0, table['head-key'])
            
        if 'head-formula' in table:
            header.set_head_formula(table['head-formula'])
            
        if 'alias' in table:
            header.set_alias(table['alias'])
                
        # record 处理        
        record_file = table.get('record_file')
        if not record_file or not os.path.isfile(record_file) or not os.access(record_file, os.R_OK):     
            table['record_file'] = os.path.join(os.path.dirname(__file__), table['record_file'])
            record_file = table['record_file']
            if not os.path.isfile(record_file) or not os.access(record_file, os.R_OK):    
                raise ValueError(f"record_file must be a readable file. Got: {record_file}")

        record = Record() 
        record.add_from_file(table['name'], record_file)
        if not record.records:
            raise ValueError(f"could not extract table record from  {record_file}, table_tag is {table['name']} ")
        
        table = Table(table['name'], table_sheet, header, record)
        table.merge_cells()
        table.set_attrs()
        
        # print(Sheet(table_sheet))
        
    
    
    # 保存工作簿
    workbook.save(excel_path)
    print(f"Excel file {excel_path} is generated by {config_file}")