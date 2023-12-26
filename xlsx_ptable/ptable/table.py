from openpyxl import Workbook
import re
import os
from .table_merge import Table_Merge
from .table_attr import Table_Attr


class Header:
    def __init__(self):
        self.headers = []
        self.aliases = {}
        self.header_col = 0
        self.record_keys = None
        self.sort_keys = None
        self.formulas = {}
        self.hash_key = None

    def add(self, header_list):
        if isinstance(header_list, list) and all(isinstance(h, str) for h in header_list):
            if self.header_col == 0:
                self.headers.append(header_list)
                self.header_col = len(header_list)
                self.record_keys = header_list
            else:
                if self.header_col != len(header_list):
                    raise ValueError("多个Header参数长度不匹配")
                self.headers.append(header_list)

        else:
            raise ValueError("Header 参数必须是字符串列表")

    def set_active(self, index, sort_keys):
        if isinstance(index, int) and 0 <= index < len(self.headers):
            self.record_keys = self.headers[index].copy()
            self.sort_keys = sort_keys
        else:
            raise ValueError("索引必须是有效的整数")


    def set_head_formula(self, head_formula):
        def __extract_patterns(input_string):
            # 使用正则表达式匹配 ${...} 格式的字符串，并提取所有匹配项
            patterns = re.findall(r'\$\{(.+?)\}', input_string)
            return patterns

        def __index_to_column(index):
            if index < 0:
                raise ValueError("Index must be a non-negative integer")
            # 将索引转换为 ASCII 值，其中 65 对应于 'A'
            column_str = ""
            while index >= 0:
                index, remainder = divmod(index, 26)
                column_str = chr(65 + remainder) + column_str
                index = index - 1
            return column_str

        def __replace_string(input_string, old_substring, new_substring):
            # 删除所有空白字符
            ss = ''.join(input_string.split())

            # 使用 str.replace 方法替换所有匹配的子字符串
            return ss.replace(old_substring, new_substring)


        # 检查 head_formula_list 是否为字典
        if not isinstance(head_formula, dict):
            raise TypeError("head_formula_list must be a dictionary")

        # 检查字典的键和值是否都是字符串
        for key, value in head_formula.items():
            if not isinstance(key, str) or not isinstance(value, str):
                raise TypeError("Both keys and values in the dictionary must be strings")

        if not self.record_keys:
            raise TypeError("set_head_formula() depends on set_active().")

        for formula_result_key, formula_str in head_formula.items():
            formula_inputs_key = __extract_patterns(formula_str)
            self.formulas[formula_result_key] = { 'inputs' : [], 'excel_formula' : None }
            forumla = self.formulas[formula_result_key]

            if formula_result_key not in self.record_keys:
                raise ValueError("formula_result_key must be in record_keys")

            for input in formula_inputs_key:
                input_formula = {'index' : None, 'key' : None, 'excel_pos' : None}
                if input not in self.record_keys:
                    raise ValueError("formula_inputs_key must be in record_keys")
                input_formula['index'] = self.record_keys.index(input)
                input_formula['key'] = input
                input_formula['excel_pos'] = __index_to_column(input_formula['index'])
                forumla['inputs'].append(input_formula)
                formula_str = __replace_string(formula_str, "${" + input + "}", input_formula['excel_pos']+"#N")

            forumla['excel_formula'] = "="  + formula_str


    def set_alias(self, alias_dict):
        if isinstance(alias_dict, dict) and all(isinstance(k, str) and isinstance(v, str) for k, v in alias_dict.items()):
            self.aliases.update(alias_dict)
        else:
            raise ValueError("字典的键和值必须是字符串")

    def set_hash_key(self, hash_key):
        if not isinstance(hash_key, list):
            raise ValueError("hash_key 必须是列表")

        for key in hash_key:
            if key not in self.record_keys:
                raise ValueError("hash_key 必须是record_keys的子集")
        self.hash_key = hash_key


    def __str__(self):
        header_str = "Headers:\n" + "\n".join(", ".join(header) for header in self.headers)
        alias_str = "Aliases:\n" + "\n".join(f"{k}: {v}" for k, v in self.aliases.items())
        active_header_str = "Active Header:\n" + ", ".join(self.record_keys) if self.record_keys else "None"
        return f"{header_str}\n\n{alias_str}\n\n{active_header_str}"




class Record:
    def __init__(self):
        self.records = []

    # 解析字符串获取类似(key@value)的字典
    @staticmethod
    def __parse_line(line, tag_name):
        if tag_name not in line:
            return {}
        pattern = r'\(([^@]+)@([^)]+)\)'
        matches = re.findall(pattern, line)

        result = {}
        for key, value in matches:
            # 尝试将 value 转换为整数
            try:
                value = int(value)
            except ValueError:
                # 如果无法转换为整数，则尝试转换为浮点数
                try:
                    value = float(value)
                except ValueError:
                    # 如果无法转换为浮点数，则保持为字符串
                    pass

            result[key] = value

        return result


    def add_from_str(self, tag_name, record_str):
        if isinstance(record_str, str):
            lines = record_str.split('\n')
            for line in lines:
                one = Record.__parse_line(line, tag_name)
                if one:
                    self.records.append(one)
        else:
            raise ValueError("参数必须是字符串")

    def add_from_file(self, tag_name, log_file):
        # 检查log_file是否为可读文件
        if not os.path.isfile(log_file) or not os.access(log_file, os.R_OK):
            raise ValueError(f"log_file must be a readable file. Got: {log_file}")

        # 打开文件并逐行读取
        with open(log_file, 'r', encoding='utf-8') as file:
            for line in file:
                one = Record.__parse_line(line, tag_name)
                if one:
                    self.records.append(one)

    def add_from_dict(self, record_dict):
        if isinstance(record_dict, dict):
            self.records.append(record_dict)
        else:
            raise ValueError("参数必须是字典")

    def __str__(self):
        records_str = "\n".join(str(record) for record in self.records)
        return f"Records:\n{records_str}"


class Table:
    def __init__(self, name_tag, worksheet, header, records):
        if not isinstance(header, Header) or not isinstance(records, list):
            raise TypeError("header and record must be Header and Record objects respectively.")
        self.__worksheet = worksheet
        self.__header = header
        self.__records = records
        self.__table_info = {'first_table': True, 'col_width': 1, 'header' : {'row_start':None, 'row_end':None}, 'record' : {'row_start':None, 'row_end':None}}
        self.__sort_key_row_index = None
        self.__name_tag = name_tag


        # replace the header name with alias name
        for h in self.__header.headers:
            for i, l in enumerate(h):
                if l in header.aliases:
                    h[i] = header.aliases[l]

        # check self.header.record_keys list strings cannot be the same
        if len(set(self.__header.record_keys)) != len(self.__header.record_keys):
            raise ValueError("用户设置header的record_keys 包含重复字符串")

        self.__to_excel()


    def __expand_and_sort_by_keys(self, records, record_keys, sort_keys):
        # 检查sort_keys是否是record_keys的子集
        if not all(key in record_keys for key in sort_keys):
            raise ValueError("sort_keys必须是record_keys的子集")

        # expand
        for index, record in enumerate(records):
            if isinstance(record, dict):
                # 检查record的所有键是否都在record_keys中
                if not all(key in record_keys for key in record.keys()):
                    raise ValueError("所有的record的keys都必须在record_keys中")

                # 检查record的必须有所有的sort_keys
                if any(not key in record.keys() for key in sort_keys):
                    raise ValueError("record的少关键key, record = {}, 关键keys = {}".format(record, sort_keys))

                records[index] = [record.get(key, None) for key in record_keys]
            else:
                raise TypeError("记录必须是字典类型")

        # sort
        # 计算sort_keys在record_keys的索引列表
        sort_indexes = [record_keys.index(key) for key in sort_keys]

        # 根据sort_indexes对self.records进行排序
        try:
            records.sort(key=lambda record: [record[index] if isinstance(record[index], (int, float)) else str(record[index]) for index in sort_indexes])
        except TypeError:
            records.sort(key=lambda record: [str(record[index]) for index in sort_indexes])


        self.__sort_key_row_index = sort_indexes


    def merge_cells(self):
        table_merge = Table_Merge(self.__worksheet, self.__table_info, self.__sort_key_row_index)
        table_merge.merge()

    def set_attrs(self):
        table_attr = Table_Attr(self.__worksheet, self.__table_info, self.__sort_key_row_index)
        table_attr.set_attr()
        pass

    def __str__(self):
        # 获取表格信息中的起始和结束行
        row_start = self.__table_info['header']['row_start']
        row_end = self.__table_info['record']['row_end']

        # 使用列表推导式构建一个包含指定行范围的字符串列表
        rows = ['\t'.join(map(str, row)) for row in self.__worksheet.iter_rows(min_row=row_start, max_row=row_end, values_only=True)]

        # 使用换行符连接所有行，构建一个完整的字符串
        return '\n'.join(rows)

    @staticmethod
    def __expand_record_by_formula(record, row_num, formulas, record_keys):
        for formula_result_key, formula in formulas.items():
            output_ind = record_keys.index(formula_result_key)

            inputs_valid = True
            for input in formula['inputs']:
                input_ind = input['index']
                # 输入为None，那么不用公式计算
                if record[input_ind] is None:
                    inputs_valid = False
                    break

                 # 支持公式嵌套，允许cell A公式的输出做为cell B的公式输入
                if isinstance(record[input_ind], (int, float)) is False:
                    if isinstance(record[input_ind], (str)) is True and not record[input_ind].startswith("="):
                        inputs_valid = False
                    break

            # 只有输入有数值、或者为公式，且输出为空，那么才用公式计算
            if inputs_valid and record[output_ind] is None:
                record[output_ind] = formula['excel_formula'].replace("#N", str(row_num))
            pass

        return record

    @staticmethod
    def __combine_two_records(record1, record2):
        for k, v in record2.items():
            if v is not None and k not in record1:
                record1[k] = v
        return record1
        pass

    def __to_excel(self):
        # 计算表格信息
        worksheet = self.__worksheet
        if worksheet.max_row == 1:
            self.__table_info['header']['row_start'] = worksheet.max_row
        else:
            # 同一个sheet中有多个table，那么table之间保留间隔行
            self.__table_info['header']['row_start'] = worksheet.max_row + 3
            worksheet.append([])
            worksheet.append([])
            self.__table_info['first_table'] = False

        self.__table_info['header']['row_end'] = self.__table_info['header']['row_start'] + len(self.__header.headers) - 1



        # 将表头写入工作表
        for header in self.__header.headers:
            worksheet.append(header)
            self.__table_info['col_width'] = max( self.__table_info['col_width'], len(header) )


        if len(self.__records) > 1 and self.__header.hash_key is None:
            raise ValueError("多个record必须设置hash_key")

        # 把多个record对象的数据合并到第一个record对象中
        def __combine_record_obj(first, second, hash_key):
            combined = first.records
            left = second.records
            for record1 in first.records:
                hash_val = hash(tuple(record1[key] for key in hash_key))
                for record2 in second.records:
                    other_hash_val = hash(tuple(record2[key] for key in hash_key))
                    if hash_val == other_hash_val:
                        record1 = Table.__combine_two_records(record1, record2)
                        left.remove(record2)
                        break
            if left:
                combined.extend(left)
            first.records = combined

        first_record_obj = self.__records[0]
        if self.__header.hash_key:
            others_record_obj = self.__records[1:]
            for other_record_obj in others_record_obj:
                __combine_record_obj(first_record_obj, other_record_obj, self.__header.hash_key)
            self.__records = [first_record_obj]
        # bugfix:  records合并后长度变化，需要重新计算record的起始和结束行
        self.__table_info['record']['row_start'] = self.__table_info['header']['row_end'] + 1
        self.__table_info['record']['row_end'] = self.__table_info['record']['row_start'] + len(self.__records[0].records) - 1

        # 以第一个记录表排序
        self.__expand_and_sort_by_keys(self.__records[0].records, self.__header.record_keys, self.__header.sort_keys)
        record_row_start = worksheet.max_row + 1

        for record in self.__records[0].records:
            record = self.__expand_record_by_formula(record, worksheet.max_row + 1, self.__header.formulas, self.__header.record_keys)
            worksheet.append(record)
        record_row_end = worksheet.max_row
