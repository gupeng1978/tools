from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font


class Table_Attr:
    def __init__(self, worksheet, table_info, sort_key_row_index):
        self.__worksheet = worksheet
        self.__table_info = table_info
        # 设置边框属性
        self.__border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        pass

    def __set_header_attr(self, row_start, row_end, col_start, col_end):
        # 设置居中对齐属性
        alignment = Alignment(horizontal='center', vertical='center')        

        # 遍历指定行范围的所有单元格        
        for row in self.__worksheet.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end):
            for cell in row:
                # 设置居中对齐属性
                cell.alignment = alignment

                # 设置边框属性
                cell.border = self.__border

                # 获取当前单元格的字体属性
                current_font = cell.font

                # 创建新的字体属性，大小比当前字体大2号，且加粗
                new_font = Font(name=current_font.name,
                                size=current_font.size + 2 if current_font.size else None,
                                bold=True)

                # 设置新的字体属性
                cell.font = new_font
                
    def __set_record_attr(self, row_start, row_end, col_start, col_end):
        
        # 设置左对齐属性
        alignment = Alignment(horizontal='left', vertical='center')
        
        # 遍历指定行范围和列范围的所有单元格
        for row in self.__worksheet.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end):
            for cell in row:
                # 设置居中对齐属性
                cell.alignment = alignment

                # 设置边框属性
                cell.border = self.__border        

    
    
    def __calculate_string_width(self, s):
        # 计算字符串的宽度，其中中文字符计为3，英文字符计为1(header里面有中文字符，且header字体比record大2号)
        return sum(3 if '\u4e00' <= char <= '\u9fff' else 1 for char in s)

    def __set_col_cell_width(self, row_start, row_end, col_start, col_end):
        # 初始化一个列表来存储每一列的最大字符宽度
        max_col_widths = [0] * self.__worksheet.max_column

        # 遍历指定行范围的所有单元格
        for row in self.__worksheet.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end, values_only=True):
            for col_index, cell_value in enumerate(row):
                # 如果单元格的值是数字，则计算其字符宽度
                if isinstance(cell_value, (int, float)):
                    # 使用自定义函数计算数字宽度
                    string_width = self.__calculate_string_width(str(cell_value))
                    # 更新该列的最大字符宽度
                    max_col_widths[col_index] = max(max_col_widths[col_index], string_width)    
                    
                # 如果单元格的值是字符串，则计算其字符宽度
                if isinstance(cell_value, str):
                    # 使用自定义函数计算字符串宽度
                    string_width = self.__calculate_string_width(cell_value)
                    # 更新该列的最大字符宽度
                    max_col_widths[col_index] = max(max_col_widths[col_index], string_width)

        # 遍历每一列，并根据最大字符宽度调整列宽
        for col_index, col_width in enumerate(max_col_widths):
            # 你可以根据需要调整这个乘数，以获得所需的列宽
            adjusted_width = col_width * 1.2
            
            # 获取列字母
            col_letter = chr(col_index + 65)
            
            # 获取当前列的宽度
            current_width = self.__worksheet.column_dimensions[col_letter].width
            
            # 如果调整的宽度大于原先单元格的宽度，则进行调整
            if adjusted_width > current_width or self.__table_info['first_table']:
                # 设置列宽
                self.__worksheet.column_dimensions[col_letter].width = adjusted_width



    def set_attr(self):        
        if 'header' in self.__table_info:
            header_info = self.__table_info['header']
            self.__set_header_attr(header_info['row_start'], header_info['row_end'], 1, self.__table_info['col_width'])
                
       
        if 'record' in self.__table_info:
            record_info = self.__table_info['record']
            self.__set_record_attr(record_info['row_start'], record_info['row_end'], 1, self.__table_info['col_width'])
            
        if 'header' in self.__table_info and 'record' in self.__table_info:
            self.__set_col_cell_width(header_info['row_start'], record_info['row_end'], 1, self.__table_info['col_width'])
            
        
 