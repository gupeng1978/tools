import unittest
from openpyxl import Workbook

from ptable import Header, Record, Table, Sheet
import os
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '../data/')


class TestTable(unittest.TestCase):
    
    def test_table_case01(self):
        header = Header()
        header.add(["model", "thread", "step","device_nn", "device_nn", "host_pcie_nn", "host_pcie_nn"])
        header.add(["model", "thread", "step","dev_ms",    "dev_fps",   "host_ms",      "host_fps"])
        header.set_active(1, ['model','thread'])
        header.set_alias({"model": "模型", 
                          "thread": "线程数", 
                          "step": "步骤", 
                          "dev_ms": "时间(ms)", 
                          "dev_fps": "帧率", 
                          "host_ms": "时间(ms)", 
                          "host_fps": "帧率",
                          "device_nn":"device推理",
                          "host_pcie_nn":"Host PCIe推理"})
        record = Record()
        
        record.add_from_dict({"model": "resnet50", 
                              "thread": '2', 
                              "step" :  "dclmdlLoadFromFile",
                              "dev_ms" : "2",
                              "host_ms" : "1.59"})

        record.add_from_dict({"model": "resnet50", 
                              "thread": '1', 
                              "step" :  "dclmdlLoadFromFile",
                              "dev_ms" : "2.59",
                              "host_ms" : "3.59"})

        record.add_from_dict({"model": "resnet50", 
                              "thread": '1', 
                              "step" :  "copyH2D",
                              "host_ms" : "0.5"})

        record.add_from_dict({"model": "resnet50", 
                              "thread": '1', 
                              "step" :  "copyD2H",
                              "host_ms" : "1.5"})

        record.add_from_dict({"model": "yolov8", 
                              "thread": '16', 
                              "step" :  "copyD2H",
                              "host_ms" : "3.5"})

        record.add_from_dict({"model": "resnet50", 
                              "thread": '2', 
                              "step" :  "copyH2D",
                              "host_ms" : "0.5"})

        record.add_from_dict({"model": "resnet50", 
                              "thread": '4', 
                              "step" :  "copyD2H",
                              "host_ms" : "3.5"})

        print(record)

         # 创建一个工作簿和工作表
        workbook = Workbook()
        worksheet = workbook.active

        table = Table(worksheet, header, record)
        table.merge_cells()
        table.set_attrs()
        
        
                
        print(table)        
        print(Sheet(worksheet,"nnp perf"))

         # 保存工作簿
        workbook.save(OUTPUT_DIR + "nnp-perf-case01.xlsx")
        
        
    def test_table_case02(self):
        header = Header()
        header.add(["$1", "input",      "input",    "output",       "output",   "$d1",      "$d1",        "$d1_copy",     "$d1_copy",       "h2d_size", "h2d_time", "ave_bw"])
        header.add(["$1", "in.format", "in.res",   "out.format",  "out.res",  "$d1.fps",  "$d1.rate",   "$d1_copy.fps", "$d1_copy.rate",  "h2d_size", "h2d_time", "ave_bw"])
        header.set_active(1, ['$1','in.format','out.format'])
        header.set_alias({"$1": "处理类别", 
                          "input": "输入", 
                          "in.format": "格式", 
                          "in.res": "分辨率", 
                          "output": "输出", 
                          "out.format": "格式", 
                          "out.res": "分辨率", 
                          
                          "$d1": "单device", 
                          "$d1.fps": "FPS", 
                          "$d1.rate": "最大rate", 
                          
                          "$d1_copy": "单device(拷贝)", 
                          "$d1_copy.fps": "FPS", 
                          "$d1_copy.rate": "最大rate", 
                          
                          
                          "h2d_size": "h2d传输大小(MB)", 
                          "h2d_time": "h2d传输时间(ms)", 
                          "ave_bw": "平均带宽(MB/S)"})
        
        
        record = Record()        
        record.add_from_dict({"$1": "Resize", 
                              "in.format": 'YUV400', 
                              "in.res": '1920x1080', 
                              "out.format": 'YUV400', 
                              "out.res": '224x224',
                              "$d1.fps" : '8109.63',
                              "$d1.rate" : '100',   
                              "$d1_copy.fps" : '2572.07', 
                              "$d1_copy.rate" : '33',  
                              "h2d_size" : '2', 
                              "h2d_time" : '2.887', 
                              "ave_bw" : '692.76'})
        record.add_from_str("""---demo : add record by logs
                            ($1@Resize),(in.format@YUV400),(in.res@640x480),(out.format@YUV400),(out.res@224x224),($d1.fps@18969.5),($d1.rate@100),($d1_copy.fps@18468),($d1_copy.rate@99),(h2d_size@0.29),(h2d_time@0.63)
                            ....
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@1920x1080),(out.format@YUV400),(out.res@224x224),  ($d1.fps@2690.19),($d1.rate@100),($d1_copy.fps@853.268),($d1_copy.rate@35),(h2d_size@5.93),(h2d_time@7.715)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@256x256),(out.format@YUV400),  (out.res@1024x1024),($d1.fps@947.92), ($d1.rate@100),($d1_copy.fps@946.962),($d1_copy.rate@100),(h2d_size@0.1875),(h2d_time@0.484)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@1920x1080),(out.format@YUV400),(out.res@1024x1024),($d1.fps@545.703),($d1.rate@100),($d1_copy.fps@503.539),($d1_copy.rate@98),(h2d_size@5.93),(h2d_time@7.683)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@1920x1080),(out.format@YUV400),(out.res@2048x2048),($d1.fps@237.727),($d1.rate@100),($d1_copy.fps@226.873),($d1_copy.rate@100),(h2d_size@5.93),(h2d_time@7.683)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@3840x2160),(out.format@YUV400),(out.res@1024x1024),($d1.fps@290.252),($d1.rate@100),($d1_copy.fps@189.045),($d1_copy.rate@72),(h2d_size@23.95),(h2d_time@29.821)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@3840x2160),(out.format@YUV400),(out.res@2048x2048),($d1.fps@139.121),($d1.rate@100),($d1_copy.fps@128.14), ($d1_copy.rate@98),(h2d_size@23.95),(h2d_time@29.821)
                            ...
                            ($1@CvtColor),(in.format@YUV420),(in.res@1920x1080),(out.format@RGB_PLANAR/RGB_PACKED/YUV444),(out.res@1920x1080),($d1.fps@480.366),($d1.rate@100),($d1_copy.fps@480.282), ($d1_copy.rate@100),(h2d_size@3),(h2d_time@4.097)
                            ($1@CvtColor),(in.format@YUV420),(in.res@1280x720), (out.format@RGB_PLANAR/RGB_PACKED/YUV444),(out.res@1280x720),($d1.fps@1082.6),($d1.rate@100),($d1_copy.fps@1080.55), ($d1_copy.rate@100),(h2d_size@1.32),(h2d_time@2.044)
                            ($1@CvtColor),(in.format@YUV420),(in.res@640x480),  (out.format@RGB_PLANAR/RGB_PACKED/YUV444),(out.res@640x480),($d1.fps@3237.86),($d1.rate@98),($d1_copy.fps@3231.45), ($d1_copy.rate@100),(h2d_size@0.44),(h2d_time@0.851)
                            """)

         # 创建一个工作簿和工作表
        workbook = Workbook()
        worksheet = workbook.active

        table = Table(worksheet, header, record)
        table.merge_cells()
        table.set_attrs()        
                       
        print(table)        
        print(Sheet(worksheet,"ive perf"))

         # 保存工作簿
        workbook.save(OUTPUT_DIR + "ive-perf-case02.xlsx")

   
    def test_table_case03(self):
        header = Header()
        header.add(["$1", "input",      "input",    "output",       "output",   "$d1",      "$d1",        "$d1_copy",     "$d1_copy",       "h2d_size", "h2d_time", "ave_bw"])
        header.add(["$1", "in.format", "in.res",   "out.format",  "out.res",  "$d1.fps",  "$d1.rate",   "$d1_copy.fps", "$d1_copy.rate",  "h2d_size", "h2d_time", "ave_bw"])
        header.set_active(1, ['$1','in.format','out.format'])
        header.set_alias({"$1": "处理类别", 
                          "input": "输入", 
                          "in.format": "格式", 
                          "in.res": "分辨率", 
                          "output": "输出", 
                          "out.format": "格式", 
                          "out.res": "分辨率", 
                          
                          "$d1": "单device", 
                          "$d1.fps": "FPS", 
                          "$d1.rate": "最大rate", 
                          
                          "$d1_copy": "单device(拷贝)", 
                          "$d1_copy.fps": "FPS", 
                          "$d1_copy.rate": "最大rate", 
                          
                          
                          "h2d_size": "h2d传输大小(MB)", 
                          "h2d_time": "h2d传输时间(ms)", 
                          "ave_bw": "平均带宽(MB/S)"})
        
        
        record = Record()        
        record.add_from_dict({"$1": "Resize", 
                              "in.format": 'YUV400', 
                              "in.res": '1920x1080', 
                              "out.format": 'YUV400', 
                              "out.res": '224x224',
                              "$d1.fps" : '8109.63',
                              "$d1.rate" : '100',   
                              "$d1_copy.fps" : '2572.07', 
                              "$d1_copy.rate" : '33',  
                              "h2d_size" : '2', 
                              "h2d_time" : '2.887', 
                              "ave_bw" : '692.76'})
        
        # 乱序排放，验证排序和合并是否准确
        record.add_from_str("""---demo : add record by logs
                            ($1@Resize),(in.format@YUV400),(in.res@640x480),(out.format@YUV400),(out.res@224x224),($d1.fps@18969.5),($d1.rate@100),($d1_copy.fps@18468),($d1_copy.rate@99),(h2d_size@0.29),(h2d_time@0.63)
                            ....
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@1920x1080),(out.format@YUV400),(out.res@224x224),  ($d1.fps@2690.19),($d1.rate@100),($d1_copy.fps@853.268),($d1_copy.rate@35),(h2d_size@5.93),(h2d_time@7.715)
                            ($1@CvtColor),(in.format@YUV420),(in.res@1280x720), (out.format@RGB_PLANAR/RGB_PACKED/YUV444),(out.res@1280x720),($d1.fps@1082.6),($d1.rate@100),($d1_copy.fps@1080.55), ($d1_copy.rate@100),(h2d_size@1.32),(h2d_time@2.044)
                            ...
                            ...
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@256x256),(out.format@YUV400),  (out.res@1024x1024),($d1.fps@947.92), ($d1.rate@100),($d1_copy.fps@946.962),($d1_copy.rate@100),(h2d_size@0.1875),(h2d_time@0.484)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@1920x1080),(out.format@YUV400),(out.res@1024x1024),($d1.fps@545.703),($d1.rate@100),($d1_copy.fps@503.539),($d1_copy.rate@98),(h2d_size@5.93),(h2d_time@7.683)
                            ...
                            ($1@CvtColor),(in.format@YUV420),(in.res@1920x1080),(out.format@RGB_PLANAR/RGB_PACKED/YUV444),(out.res@1920x1080),($d1.fps@480.366),($d1.rate@100),($d1_copy.fps@480.282), ($d1_copy.rate@100),(h2d_size@3),(h2d_time@4.097)
                            ...
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@1920x1080),(out.format@YUV400),(out.res@2048x2048),($d1.fps@237.727),($d1.rate@100),($d1_copy.fps@226.873),($d1_copy.rate@100),(h2d_size@5.93),(h2d_time@7.683)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@3840x2160),(out.format@YUV400),(out.res@1024x1024),($d1.fps@290.252),($d1.rate@100),($d1_copy.fps@189.045),($d1_copy.rate@72),(h2d_size@23.95),(h2d_time@29.821)
                            ($1@Resize),(in.format@RGB_PLANAR),(in.res@3840x2160),(out.format@YUV400),(out.res@2048x2048),($d1.fps@139.121),($d1.rate@100),($d1_copy.fps@128.14), ($d1_copy.rate@98),(h2d_size@23.95),(h2d_time@29.821)
                            ...
                            ($1@CvtColor),(in.format@YUV420),(in.res@640x480),  (out.format@RGB_PLANAR/RGB_PACKED/YUV444),(out.res@640x480),($d1.fps@3237.86),($d1.rate@98),($d1_copy.fps@3231.45), ($d1_copy.rate@100),(h2d_size@0.44),(h2d_time@0.851)
                            """)

         # 创建一个工作簿和工作表
        workbook = Workbook()
        worksheet = workbook.active

        table = Table(worksheet, header, record)
        table.merge_cells()
        table.set_attrs()        
                       
        print(table)        
        print(Sheet(worksheet,"ive perf"))

         # 保存工作簿
        workbook.save(OUTPUT_DIR + "ive-perf-case03.xlsx")

# 如果是直接运行这个文件，那么执行测试
if __name__ == '__main__':
    unittest.main()
