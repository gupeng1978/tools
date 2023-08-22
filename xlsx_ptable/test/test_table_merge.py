import unittest
from openpyxl import Workbook
from table import Table_Merge, Sheet

class TestTableMerge(unittest.TestCase):
    
    def test_merge_header_case01(self):
        ws = Workbook().active
        sheet = Sheet(ws)
        ws.append(["A"])
        ws.append(["A"])
        table_info = {'header': {'row_start':1, 'row_end' : 2},}
        sort_key_row_index = [0,] # 根据A1排序并合并
        table_merge = Table_Merge(ws, table_info, sort_key_row_index)
        
        table_merge.merge()
        print(sheet)
        
        
        # 使用断言方法验证合并后的单元格
        self.assertEqual(len(ws.merged_cells.ranges), 1)
        for merged_cell in ws.merged_cells.ranges:
            if 'A1' in merged_cell:
                self.assertEqual(merged_cell.coord, "A1:A2")

        
        

    def test_merge_header_case02(self):
        ws = Workbook().active
        sheet = Sheet(ws)
        ws.append(["A",     "A",    "B",    "C",    "C"])
        ws.append(["A1",    "A2",   "B",   "C1",   "C2"])
        table_info = {'header': {'row_start':1, 'row_end' : 2},}
        sort_key_row_index = [0,] # 根据A1排序并合并
        table_merge = Table_Merge(ws, table_info, sort_key_row_index)
        
        table_merge.merge()
        print(sheet)
        
        
        # 使用断言方法验证合并后的单元格
        self.assertEqual(len(ws.merged_cells.ranges), 3)
        for merged_cell in ws.merged_cells.ranges:
            if 'A1' in merged_cell:
                self.assertEqual(merged_cell.coord, "A1:B1")
            if 'D1' in merged_cell:
                self.assertEqual(merged_cell.coord, "D1:E1")
            if 'C1' in merged_cell:
                self.assertEqual(merged_cell.coord, "C1:C2")
        
        
    def test_merge_record_case01(self):
        ws = Workbook().active
        sheet = Sheet(ws)
        ws.append(["A",    "B",   "C",    "D",    "E"])
        ws.append(["A",    "B",   "C",   "D1",   "E1"])
        ws.append(["A",    "B1",   "C",   "D1",   "E2"])
        table_info = {'record': {'row_start':1, 'row_end' : 3},}
        sort_key_row_index = [0, 1] # 根据A,B排序并合并
        table_merge = Table_Merge(ws, table_info, sort_key_row_index)
        table_merge.merge()
        print(sheet)
        
        #仅允许A，B列merge
        self.assertEqual(len(ws.merged_cells.ranges), 2)
        for merged_cell in ws.merged_cells.ranges:
            if 'A1' in merged_cell:
                self.assertEqual(merged_cell.coord, "A1:A3")
            if 'B1' in merged_cell:
                self.assertEqual(merged_cell.coord, "B1:B2")
       
        
        
    def test_merge_record_case02(self):
        ws = Workbook().active
        sheet = Sheet(ws)
        ws.append(["A1",    "B",   "C",    "D",    "E"])
        ws.append(["A1",    "B",   "C",   "D1",   "E1"])
        ws.append(["A1",    "B1",   "C",   "D1",   "E2"])
        ws.append(["A2",    "B1",   "C",   "D1",   "E2"])
        ws.append(["A2",    "B1",   "C",   "D1",   "E2"])
        table_info = {'record': {'row_start':1, 'row_end' : 5},}
        sort_key_row_index = [0, 1, 2] # 根据A,B排序并合并
        table_merge = Table_Merge(ws, table_info, sort_key_row_index)
        table_merge.merge()
        print(sheet)
        
        #仅允许A，B,C 列merge， B的合并受A约束， C的合并受B约束
        self.assertEqual(len(ws.merged_cells.ranges), 6)
        for merged_cell in ws.merged_cells.ranges:
            if 'A1' in merged_cell:
                self.assertEqual(merged_cell.coord, "A1:A3")
            if 'A4' in merged_cell:
                self.assertEqual(merged_cell.coord, "A4:A5")
            if 'B1' in merged_cell:
                self.assertEqual(merged_cell.coord, "B1:B2")
            if 'B4' in merged_cell:
                self.assertEqual(merged_cell.coord, "B4:B5")
            if 'C1' in merged_cell:
                self.assertEqual(merged_cell.coord, "C1:C2")
            if 'C4' in merged_cell:
                self.assertEqual(merged_cell.coord, "C4:C5")
         
    

# 如果是直接运行这个文件，那么执行测试
if __name__ == '__main__':
    unittest.main()
